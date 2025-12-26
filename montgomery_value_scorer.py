#!/usr/bin/env python3
"""
Value Investing Stock Scorer - Based on Roger Montgomery's Value.able methodology
Evaluates companies based on quality metrics with emphasis on ROE, debt, and cash flow

NOTE: This script requires an active internet connection to fetch data from Yahoo Finance.
      If running in a restricted environment, you may need to configure proxy settings.
"""

import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import sys
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import warnings
warnings.filterwarnings('ignore')

class MontgomeryValueScorer:
    """
    Implements Roger Montgomery's value investing scoring methodology
    Focus on high ROE, low debt, consistent earnings, and attractive valuations
    """
    
    def __init__(self):
        self.scores = {}
        self.raw_data = {}
        self.existing_results = None

    def load_existing_results(self, output_file):
        """Load existing results from Excel file for resume functionality"""
        try:
            import os
            if not os.path.exists(output_file):
                print(f"No existing file found at {output_file}. Starting fresh analysis.")
                return None

            # Read existing data (header is at row 3, 0-indexed)
            df = pd.read_excel(output_file, sheet_name='Value Investment Analysis', header=3)

            if df.empty or 'Ticker' not in df.columns:
                print(f"Existing file found but no valid data. Starting fresh analysis.")
                return None

            print(f"Found existing results with {len(df)} tickers already processed.")
            print(f"Last processed ticker: {df['Ticker'].iloc[-1]}")
            self.existing_results = df
            return df

        except Exception as e:
            print(f"Could not load existing results: {e}")
            print("Starting fresh analysis.")
            return None

    def fetch_stock_data(self, ticker):
        """Fetch comprehensive financial data for a stock"""
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            
            # Get financial statements
            income_stmt = stock.income_stmt
            balance_sheet = stock.balance_sheet
            cash_flow = stock.cash_flow
            
            # Historical data for volatility and consistency
            hist = stock.history(period="5y")
            
            return {
                'ticker': ticker,
                'info': info,
                'income_stmt': income_stmt,
                'balance_sheet': balance_sheet,
                'cash_flow': cash_flow,
                'history': hist
            }
        except Exception as e:
            print(f"Error fetching {ticker}: {e}")
            return None
    
    def calculate_intrinsic_value_dcf(self, data, metrics):
        """
        Calculate intrinsic value using Discounted Cash Flow method
        Uses Free Cash Flow with conservative assumptions
        """
        try:
            fcf = metrics.get('Free Cash Flow (B)', 0) * 1e9  # Convert back to actual value
            if fcf <= 0:
                return None

            # Quality gate: Skip DCF for very low quality companies
            roe = metrics.get('ROE (%)', 0)
            if roe < -20:  # Severely negative ROE indicates distressed company
                return None

            # Get growth rates
            earnings_growth = metrics.get('Earnings Growth (%)', 0) / 100
            revenue_growth = metrics.get('Revenue Growth (%)', 0) / 100
            revenue_cagr = metrics.get('Revenue CAGR 3Y (%)', 0) / 100

            # FCF growth should be constrained by revenue growth
            # A company can't sustainably grow FCF faster than revenue long-term
            # Use the more conservative of revenue metrics
            revenue_constraint = min(revenue_growth, revenue_cagr) if revenue_cagr != 0 else revenue_growth

            # For declining revenue companies, use the decline rate (don't assume growth)
            if revenue_constraint < 0:
                # Revenue is declining - FCF will likely follow
                growth_rate = max(revenue_constraint, -0.10)  # Cap decline at -10%
            else:
                # Revenue is growing - use minimum of earnings growth and revenue growth
                growth_rate = min(earnings_growth, revenue_constraint, 0.15)
                if growth_rate <= 0:
                    growth_rate = min(revenue_constraint, 0.05)  # Default to revenue growth or 5%

            # Discount rate (WACC approximation)
            # Base rate on debt levels AND profitability
            debt_to_equity = metrics.get('Debt to Equity', 0)

            # Start with debt-based rate
            if debt_to_equity < 0.3:
                discount_rate = 0.10  # Low debt risk
            elif debt_to_equity < 1.0:
                discount_rate = 0.12  # Moderate debt risk
            elif debt_to_equity < 2.0:
                discount_rate = 0.15  # High debt risk
            else:
                discount_rate = 0.18  # Very high debt risk

            # Adjust for profitability risk
            if roe < 0:
                discount_rate += 0.03  # Add 3% for unprofitable companies
            elif roe < 10:
                discount_rate += 0.01  # Add 1% for low profitability

            # Adjust for declining revenue
            if revenue_constraint < -0.05:
                discount_rate += 0.02  # Add 2% for significant revenue decline

            # Terminal growth rate - should be conservative and below discount rate
            # For declining companies, terminal growth should be lower or zero
            if revenue_constraint < 0:
                terminal_growth = 0.01  # 1% terminal growth for declining companies
            else:
                terminal_growth = 0.025  # 2.5% for stable/growing companies

            # Project FCF for 5 years
            projected_fcf = []
            current_fcf = fcf
            for year in range(1, 6):
                current_fcf = current_fcf * (1 + growth_rate)
                discounted_fcf = current_fcf / ((1 + discount_rate) ** year)
                projected_fcf.append(discounted_fcf)

            # Terminal value using perpetuity growth model
            terminal_fcf = current_fcf * (1 + terminal_growth)
            # Guard against division by zero or invalid assumptions
            if discount_rate <= terminal_growth:
                return None
            terminal_value = terminal_fcf / (discount_rate - terminal_growth)
            discounted_terminal_value = terminal_value / ((1 + discount_rate) ** 5)

            # Enterprise value = sum of projected FCF + terminal value
            enterprise_value = sum(projected_fcf) + discounted_terminal_value

            # Get shares outstanding
            shares_outstanding = data['info'].get('sharesOutstanding', 0)
            if shares_outstanding == 0:
                return None

            # Adjust for net cash/debt
            total_cash = metrics.get('Total Cash (B)', 0) * 1e9
            total_debt = data['info'].get('totalDebt', 0)
            net_cash = total_cash - total_debt

            # Equity value = Enterprise value + net cash
            equity_value = enterprise_value + net_cash

            # Intrinsic value per share
            intrinsic_value = equity_value / shares_outstanding

            return intrinsic_value if intrinsic_value > 0 else None

        except Exception as e:
            return None

    def calculate_intrinsic_value_graham(self, metrics):
        """
        Calculate intrinsic value using Benjamin Graham Formula
        Formula: V = EPS × (8.5 + 2g) × 4.4 / Y
        Where: EPS = earnings per share, g = expected growth rate, Y = current AAA bond yield
        Simplified version: V = EPS × (8.5 + 2g)
        """
        try:
            # Get EPS (trailing)
            eps = metrics.get('info_eps', 0)
            if eps <= 0:
                return None

            # Get growth rates
            earnings_growth = metrics.get('Earnings Growth (%)', 0)
            revenue_growth = metrics.get('Revenue Growth (%)', 0)
            revenue_cagr = metrics.get('Revenue CAGR 3Y (%)', 0)

            # Use the more conservative revenue metric
            revenue_constraint = min(revenue_growth, revenue_cagr) if revenue_cagr != 0 else revenue_growth

            # For declining revenue companies, growth should be constrained
            if revenue_constraint < 0:
                # Declining revenue - use 0 growth (Graham's base case)
                growth_rate = 0
            else:
                # Growing revenue - use minimum of earnings and revenue growth
                growth_rate = min(earnings_growth, revenue_constraint)
                # Cap between 0-15% for conservatism
                growth_rate = max(0, min(growth_rate, 15))

            # Graham formula (simplified without bond yield adjustment)
            intrinsic_value = eps * (8.5 + 2 * growth_rate)

            return intrinsic_value if intrinsic_value > 0 else None

        except Exception as e:
            return None

    def calculate_metrics(self, data):
        """Calculate key financial metrics for scoring"""
        if not data:
            return None

        info = data['info']
        ticker = data['ticker']

        metrics = {'Ticker': ticker}
        
        # Company info
        metrics['Company Name'] = info.get('longName', 'N/A')
        metrics['Sector'] = info.get('sector', 'N/A')
        metrics['Industry'] = info.get('industry', 'N/A')
        
        # Price and market data
        metrics['Current Price'] = info.get('currentPrice', info.get('regularMarketPrice', 0))
        metrics['Market Cap (B)'] = info.get('marketCap', 0) / 1e9
        
        # Montgomery's key metric: Return on Equity (ROE)
        metrics['ROE (%)'] = info.get('returnOnEquity', 0) * 100 if info.get('returnOnEquity') else 0
        
        # Debt metrics - Montgomery prefers low/no debt companies
        total_debt = info.get('totalDebt', 0)
        # Get equity from balance sheet (more reliable than info dict)
        total_equity = 0
        if not data['balance_sheet'].empty and 'Stockholders Equity' in data['balance_sheet'].index:
            total_equity = data['balance_sheet'].loc['Stockholders Equity'].iloc[0]
            if pd.isna(total_equity):
                total_equity = 0

        # Handle missing or negative equity properly
        if total_equity <= 0:
            # No valid equity data or negative equity - indicates problematic financials
            metrics['Debt to Equity'] = 999
        else:
            metrics['Debt to Equity'] = total_debt / total_equity
        
        # Cash and liquidity
        metrics['Total Cash (B)'] = info.get('totalCash', 0) / 1e9

        # Get Total Revenue from income statement
        total_revenue = 0
        if not data['income_stmt'].empty and 'Total Revenue' in data['income_stmt'].index:
            try:
                total_revenue = data['income_stmt'].loc['Total Revenue'].iloc[0]
                if pd.isna(total_revenue):
                    total_revenue = 0
            except (KeyError, IndexError, TypeError):
                total_revenue = 0
        metrics['Total Revenue (B)'] = total_revenue / 1e9 if total_revenue else 0

        # Get Free Cash Flow - try info first, then cash flow statement
        fcf = info.get('freeCashflow', 0)
        if fcf == 0 and not data['cash_flow'].empty:
            try:
                if 'Free Cash Flow' in data['cash_flow'].index:
                    fcf = data['cash_flow'].loc['Free Cash Flow'].iloc[0]
                    if pd.isna(fcf):
                        fcf = 0
            except (KeyError, IndexError, TypeError):
                pass
        metrics['Free Cash Flow (B)'] = fcf / 1e9
        
        # Profitability metrics
        # Note: netMargins might not be available in newer yfinance versions
        net_margin = info.get('netMargins') or info.get('profitMargins')

        # If not in info, try to calculate from financial statements
        if not net_margin and not data['income_stmt'].empty:
            try:
                if 'Net Income' in data['income_stmt'].index and 'Total Revenue' in data['income_stmt'].index:
                    net_income = data['income_stmt'].loc['Net Income'].iloc[0]
                    revenue = data['income_stmt'].loc['Total Revenue'].iloc[0]
                    if revenue and revenue > 0 and not pd.isna(net_income) and not pd.isna(revenue):
                        net_margin = net_income / revenue
            except (KeyError, IndexError, TypeError):
                pass

        metrics['Net Margin (%)'] = net_margin * 100 if net_margin else 0
        metrics['Operating Margin (%)'] = info.get('operatingMargins', 0) * 100 if info.get('operatingMargins') else 0
        metrics['ROA (%)'] = info.get('returnOnAssets', 0) * 100 if info.get('returnOnAssets') else 0
        
        # Growth metrics - Calculate from income statement for accuracy
        # Revenue growth from annual income statement (not quarterly yfinance metric)
        revenue_growth_annual = 0
        revenue_cagr_3y = 0
        revenue_growth_consistency = 0

        if not data['income_stmt'].empty and 'Total Revenue' in data['income_stmt'].index:
            try:
                revenues = data['income_stmt'].loc['Total Revenue'].dropna()
                if len(revenues) >= 2:
                    # Most recent YoY growth (year 0 vs year 1)
                    rev_current = revenues.iloc[0]
                    rev_prior = revenues.iloc[1]
                    if rev_prior > 0:
                        revenue_growth_annual = ((rev_current - rev_prior) / rev_prior) * 100

                    # 3-year CAGR if we have enough data
                    if len(revenues) >= 4:
                        rev_3y_ago = revenues.iloc[3]
                        if rev_3y_ago > 0 and rev_current > 0:
                            revenue_cagr_3y = ((rev_current / rev_3y_ago) ** (1/3) - 1) * 100
                    elif len(revenues) >= 3:
                        # Use 2-year CAGR if only 3 years available
                        rev_2y_ago = revenues.iloc[2]
                        if rev_2y_ago > 0 and rev_current > 0:
                            revenue_cagr_3y = ((rev_current / rev_2y_ago) ** (1/2) - 1) * 100

                    # Revenue growth consistency (std dev of annual growth rates)
                    if len(revenues) >= 3:
                        growth_rates = []
                        for i in range(len(revenues) - 1):
                            if revenues.iloc[i+1] > 0:
                                gr = ((revenues.iloc[i] - revenues.iloc[i+1]) / revenues.iloc[i+1]) * 100
                                growth_rates.append(gr)
                        if growth_rates:
                            revenue_growth_consistency = np.std(growth_rates)
            except (KeyError, IndexError, TypeError, ValueError):
                pass

        metrics['Revenue Growth (%)'] = revenue_growth_annual
        metrics['Revenue CAGR 3Y (%)'] = revenue_cagr_3y
        metrics['Revenue Growth Consistency'] = revenue_growth_consistency
        metrics['Earnings Growth (%)'] = info.get('earningsGrowth', 0) * 100 if info.get('earningsGrowth') else 0
        
        # Valuation metrics
        metrics['P/E Ratio'] = info.get('trailingPE', 0)
        metrics['Forward P/E'] = info.get('forwardPE', 0)
        metrics['P/B Ratio'] = info.get('priceToBook', 0)
        metrics['PEG Ratio'] = info.get('trailingPegRatio', 0) if info.get('trailingPegRatio') else 0
        metrics['EV/EBITDA'] = info.get('enterpriseToEbitda', 0)
        
        # Dividend metrics
        # Note: yfinance returns dividendYield already as percentage value (0.75 = 0.75%, 2.93 = 2.93%)
        metrics['Dividend Yield (%)'] = info.get('dividendYield', 0) if info.get('dividendYield') else 0
        metrics['Payout Ratio (%)'] = info.get('payoutRatio', 0) * 100 if info.get('payoutRatio') else 0
        
        # Additional quality metrics
        metrics['Current Ratio'] = info.get('currentRatio', 0)
        metrics['Quick Ratio'] = info.get('quickRatio', 0)

        # EPS for intrinsic value calculations
        metrics['info_eps'] = info.get('trailingEps', 0)

        # Calculate 5-year average ROE if possible
        try:
            if not data['balance_sheet'].empty and not data['income_stmt'].empty:
                net_income = data['income_stmt'].loc['Net Income'] if 'Net Income' in data['income_stmt'].index else None
                equity = data['balance_sheet'].loc['Stockholders Equity'] if 'Stockholders Equity' in data['balance_sheet'].index else None
                
                if net_income is not None and equity is not None:
                    roe_list = []
                    for col in net_income.index[:min(5, len(net_income.index))]:
                        if col in equity.index:
                            ni = net_income[col]
                            eq = equity[col]
                            if eq > 0 and not pd.isna(ni) and not pd.isna(eq):
                                roe_list.append((ni / eq) * 100)
                    
                    if roe_list:
                        metrics['5Y Avg ROE (%)'] = np.mean(roe_list)
                        metrics['ROE Consistency'] = np.std(roe_list) if len(roe_list) > 1 else 0
                    else:
                        metrics['5Y Avg ROE (%)'] = metrics['ROE (%)']
                        metrics['ROE Consistency'] = 0
                else:
                    metrics['5Y Avg ROE (%)'] = metrics['ROE (%)']
                    metrics['ROE Consistency'] = 0
            else:
                metrics['5Y Avg ROE (%)'] = metrics['ROE (%)']
                metrics['ROE Consistency'] = 0
        except (KeyError, IndexError, TypeError, ValueError):
            metrics['5Y Avg ROE (%)'] = metrics['ROE (%)']
            metrics['ROE Consistency'] = 0

        # Calculate intrinsic values
        current_price = metrics.get('Current Price', 0)

        # DCF-based intrinsic value
        iv_dcf = self.calculate_intrinsic_value_dcf(data, metrics)
        metrics['Intrinsic Value (DCF)'] = iv_dcf if iv_dcf else 0

        # Graham formula intrinsic value
        iv_graham = self.calculate_intrinsic_value_graham(metrics)
        metrics['Intrinsic Value (Graham)'] = iv_graham if iv_graham else 0

        # Calculate average intrinsic value (use available methods)
        iv_values = [v for v in [iv_dcf, iv_graham] if v and v > 0]
        avg_iv = np.mean(iv_values) if iv_values else 0
        metrics['Avg Intrinsic Value'] = avg_iv

        # Calculate upside/downside potential
        if current_price > 0 and avg_iv > 0:
            upside = ((avg_iv - current_price) / current_price) * 100
            metrics['Upside Potential (%)'] = upside
            metrics['Margin of Safety (%)'] = upside  # Positive = undervalued, Negative = overvalued
        else:
            metrics['Upside Potential (%)'] = 0
            metrics['Margin of Safety (%)'] = 0

        # Value recommendation based on upside and quality
        if avg_iv > 0 and current_price > 0:
            upside_pct = metrics['Upside Potential (%)']
            if upside_pct >= 50:
                metrics['Value Rating'] = 'Strong Buy'
            elif upside_pct >= 25:
                metrics['Value Rating'] = 'Buy'
            elif upside_pct >= 10:
                metrics['Value Rating'] = 'Hold'
            elif upside_pct >= -10:
                metrics['Value Rating'] = 'Fair Value'
            else:
                metrics['Value Rating'] = 'Overvalued'
        else:
            metrics['Value Rating'] = 'N/A'

        # Filter stocks with insufficient revenue or cashflow
        # Minimum thresholds: $10M revenue, $1M free cash flow
        revenue_millions = metrics.get('Total Revenue (B)', 0) * 1000  # Convert billions to millions
        fcf_millions = metrics.get('Free Cash Flow (B)', 0) * 1000  # Convert billions to millions

        if revenue_millions < 10:
            print(f"  Filtering {ticker}: Revenue ${revenue_millions:.2f}M < $10M minimum")
            return None

        if fcf_millions < 1:
            print(f"  Filtering {ticker}: Free Cash Flow ${fcf_millions:.2f}M < $1M minimum")
            return None

        return metrics
    
    def calculate_montgomery_score(self, metrics):
        """
        Calculate Montgomery-style value score (0-100)
        Heavy emphasis on ROE, debt levels, and cash generation
        """
        if not metrics:
            return 0
        
        score = 0
        max_score = 100
        
        # 1. ROE Score (30 points) - Montgomery's primary quality metric
        roe = metrics.get('ROE (%)', 0)
        avg_roe = metrics.get('5Y Avg ROE (%)', roe)
        
        if avg_roe >= 20:
            roe_score = 30
        elif avg_roe >= 15:
            roe_score = 25
        elif avg_roe >= 10:
            roe_score = 15
        elif avg_roe >= 5:
            roe_score = 5
        else:
            roe_score = 0
        
        # Bonus for consistency
        consistency = metrics.get('ROE Consistency', 999)
        if consistency < 5 and avg_roe > 10:  # Low standard deviation is good
            roe_score = min(30, roe_score + 5)
        
        score += roe_score
        
        # 2. Debt Score (25 points) - Montgomery prefers low/no debt
        debt_to_equity = metrics.get('Debt to Equity', 999)
        
        if debt_to_equity == 0:
            debt_score = 25
        elif debt_to_equity < 0.1:
            debt_score = 23
        elif debt_to_equity < 0.3:
            debt_score = 20
        elif debt_to_equity < 0.5:
            debt_score = 15
        elif debt_to_equity < 1.0:
            debt_score = 8
        else:
            debt_score = 0
        
        score += debt_score
        
        # 3. Free Cash Flow Score (20 points)
        fcf = metrics.get('Free Cash Flow (B)', 0)
        market_cap = metrics.get('Market Cap (B)', 1)
        fcf_yield = (fcf / market_cap * 100) if market_cap > 0 else 0
        
        if fcf_yield >= 10:
            fcf_score = 20
        elif fcf_yield >= 7:
            fcf_score = 16
        elif fcf_yield >= 5:
            fcf_score = 12
        elif fcf_yield >= 3:
            fcf_score = 8
        elif fcf > 0:
            fcf_score = 4
        else:
            fcf_score = 0
        
        score += fcf_score
        
        # 4. Valuation Score (15 points)
        pe = metrics.get('P/E Ratio', 999)
        pb = metrics.get('P/B Ratio', 999)
        
        # P/E component
        if 0 < pe <= 10:
            pe_score = 8
        elif 10 < pe <= 15:
            pe_score = 6
        elif 15 < pe <= 20:
            pe_score = 4
        elif 20 < pe <= 25:
            pe_score = 2
        else:
            pe_score = 0
        
        # P/B relative to ROE (Montgomery's value approach)
        if roe > 0 and pb > 0:
            pb_roe_ratio = pb / (roe / 10)  # Normalize ROE to compare with P/B
            if pb_roe_ratio < 1.5:
                pb_score = 7
            elif pb_roe_ratio < 2.0:
                pb_score = 5
            elif pb_roe_ratio < 3.0:
                pb_score = 3
            else:
                pb_score = 0
        else:
            pb_score = 0
        
        score += min(15, pe_score + pb_score)
        
        # 5. Profitability Score (10 points)
        net_margin = metrics.get('Net Margin (%)', 0)
        op_margin = metrics.get('Operating Margin (%)', 0)
        
        if net_margin >= 20:
            margin_score = 5
        elif net_margin >= 15:
            margin_score = 4
        elif net_margin >= 10:
            margin_score = 3
        elif net_margin >= 5:
            margin_score = 1
        else:
            margin_score = 0
        
        if op_margin >= 25:
            margin_score += 5
        elif op_margin >= 20:
            margin_score += 4
        elif op_margin >= 15:
            margin_score += 3
        elif op_margin >= 10:
            margin_score += 1
        
        score += min(10, margin_score)
        
        return min(100, score)

    def _process_single_ticker(self, ticker):
        """Process a single ticker and return its metrics (helper for parallel processing)"""
        ticker = ticker.strip().upper()
        if not ticker:
            return None

        print(f"Processing {ticker}...")
        data = self.fetch_stock_data(ticker)

        if data:
            metrics = self.calculate_metrics(data)
            if metrics:
                score = self.calculate_montgomery_score(metrics)
                metrics['Montgomery Score'] = score

                # Determine quality rating based on score
                if score >= 80:
                    metrics['Quality Rating'] = 'A+'
                elif score >= 70:
                    metrics['Quality Rating'] = 'A'
                elif score >= 60:
                    metrics['Quality Rating'] = 'B+'
                elif score >= 50:
                    metrics['Quality Rating'] = 'B'
                elif score >= 40:
                    metrics['Quality Rating'] = 'C+'
                elif score >= 30:
                    metrics['Quality Rating'] = 'C'
                else:
                    metrics['Quality Rating'] = 'D'

                return metrics

        return None

    def process_tickers(self, tickers, skip_processed=True, workers=1):
        """Process a list of tickers and calculate scores with optional parallel processing"""
        results = []

        # Get list of already processed tickers if resuming
        processed_tickers = set()
        if skip_processed and self.existing_results is not None:
            processed_tickers = set(self.existing_results['Ticker'].str.upper())
            print(f"Skipping {len(processed_tickers)} already processed tickers...")

        # Filter out already processed tickers
        tickers_to_process = []
        for ticker in tickers:
            ticker = ticker.strip().upper()
            if not ticker:
                continue

            # Skip if already processed
            if ticker in processed_tickers:
                print(f"Skipping {ticker} (already processed)")
                continue

            tickers_to_process.append(ticker)

        if not tickers_to_process:
            print("No new tickers to process")
            return pd.DataFrame()

        print(f"\nProcessing {len(tickers_to_process)} tickers with {workers} worker(s)...")

        # Process tickers
        if workers == 1:
            # Sequential processing (original behavior)
            for ticker in tickers_to_process:
                metrics = self._process_single_ticker(ticker)
                if metrics:
                    results.append(metrics)
        else:
            # Parallel processing using ThreadPoolExecutor
            with ThreadPoolExecutor(max_workers=workers) as executor:
                # Submit all tasks
                future_to_ticker = {executor.submit(self._process_single_ticker, ticker): ticker
                                   for ticker in tickers_to_process}

                # Collect results as they complete
                for future in as_completed(future_to_ticker):
                    ticker = future_to_ticker[future]
                    try:
                        metrics = future.result()
                        if metrics:
                            results.append(metrics)
                    except Exception as e:
                        print(f"Error processing {ticker}: {e}")

        return pd.DataFrame(results)
    
    def create_excel_report(self, df, output_file):
        """Create formatted Excel report with color coding"""

        # Merge with existing results if resuming
        if self.existing_results is not None and not self.existing_results.empty:
            print(f"Merging {len(df)} new results with {len(self.existing_results)} existing results...")
            df = pd.concat([self.existing_results, df], ignore_index=True)

        # Sort by Montgomery Score (important after merging)
        if 'Montgomery Score' in df.columns:
            df = df.sort_values('Montgomery Score', ascending=False)

        wb = Workbook()
        ws = wb.active
        ws.title = "Value Investment Analysis"
        
        # Add title
        ws['A1'] = "Montgomery Value Investment Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        # Add headers starting from row 4
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data
        for row_idx, row in enumerate(df.itertuples(index=False), 5):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Format numbers
                header = headers[col_idx - 1]
                if isinstance(value, (int, float)) and not pd.isna(value):
                    if 'Price' in header or '$' in header:
                        cell.number_format = '$#,##0.00'
                    elif '%' in header or 'Ratio' in header or 'Score' in header:
                        cell.number_format = '0.00'
                    elif header in ['Market Cap (B)', 'Total Cash (B)', 'Free Cash Flow (B)']:
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '#,##0.00'
                
                # Apply conditional formatting for Montgomery Score
                if header == 'Montgomery Score':
                    if value >= 70:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    elif value >= 50:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(color="9C5700")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
                
                # Apply conditional formatting for Quality Rating
                if header == 'Quality Rating':
                    if value in ['A+', 'A']:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    elif value in ['B+', 'B']:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(color="9C5700")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
                
                # Highlight exceptional ROE
                if header == 'ROE (%)' and isinstance(value, (int, float)):
                    if value >= 20:
                        cell.font = Font(bold=True, color="006100")
                
                # Highlight low debt
                if header == 'Debt to Equity' and isinstance(value, (int, float)):
                    if value < 0.3:
                        cell.font = Font(bold=True, color="006100")
                    elif value > 1:
                        cell.font = Font(color="9C0006")

                # Highlight high upside potential
                if header == 'Upside Potential (%)' and isinstance(value, (int, float)):
                    if value >= 50:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    elif value >= 25:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(bold=True, color="9C5700")
                    elif value < -20:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")

                # Highlight margin of safety
                if header == 'Margin of Safety (%)' and isinstance(value, (int, float)):
                    if value >= 30:
                        cell.font = Font(bold=True, color="006100")
                    elif value < 0:
                        cell.font = Font(color="9C0006")

                # Color code value rating
                if header == 'Value Rating':
                    if value == 'Strong Buy':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    elif value == 'Buy':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(bold=True, color="9C5700")
                    elif value == 'Overvalued':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")

                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Add summary sheet
        ws2 = wb.create_sheet("Scoring Methodology")
        
        methodology = [
            ["Montgomery Value Scoring Methodology", "", ""],
            ["", "", ""],
            ["Component", "Weight", "Criteria"],
            ["Return on Equity (ROE)", "30%", "≥20%: 30pts, ≥15%: 25pts, ≥10%: 15pts"],
            ["Debt Levels", "25%", "D/E=0: 25pts, <0.3: 20pts, <0.5: 15pts"],
            ["Free Cash Flow", "20%", "FCF Yield ≥10%: 20pts, ≥7%: 16pts, ≥5%: 12pts"],
            ["Valuation", "15%", "P/E and P/B relative to ROE"],
            ["Profitability", "10%", "Net Margin and Operating Margin"],
            ["", "", ""],
            ["Quality Ratings", "", ""],
            ["A+", "80-100", "Exceptional quality, strong moat"],
            ["A", "70-79", "High quality, competitive advantages"],
            ["B+", "60-69", "Good quality, above average"],
            ["B", "50-59", "Average quality, acceptable"],
            ["C+", "40-49", "Below average, some concerns"],
            ["C", "30-39", "Poor quality, significant issues"],
            ["D", "0-29", "Very poor quality, avoid"],
            ["", "", ""],
            ["Intrinsic Value Methods", "", ""],
            ["DCF Method", "", "5-year FCF projection with terminal value"],
            ["", "", "Conservative growth assumptions (max 15%)"],
            ["", "", "Risk-adjusted discount rate (10-15% based on debt)"],
            ["Graham Formula", "", "V = EPS × (8.5 + 2g), where g is growth rate"],
            ["Avg Intrinsic Value", "", "Average of DCF and Graham methods"],
            ["", "", ""],
            ["Value Ratings", "", ""],
            ["Strong Buy", "≥50% upside", "Significantly undervalued"],
            ["Buy", "25-50% upside", "Attractive value opportunity"],
            ["Hold", "10-25% upside", "Moderate upside potential"],
            ["Fair Value", "-10% to 10%", "Fairly priced"],
            ["Overvalued", "<-10%", "Trading above intrinsic value"],
            ["", "", ""],
            ["Key Principles", "", ""],
            ["1. High ROE with consistency", "", ""],
            ["2. Low or no debt preferred", "", ""],
            ["3. Strong free cash flow generation", "", ""],
            ["4. Reasonable valuation relative to quality", "", ""],
            ["5. Sustainable competitive advantages", "", ""],
            ["6. Margin of safety: Buy when price < intrinsic value", "", ""],
        ]
        
        for row_idx, row_data in enumerate(methodology, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(size=14, bold=True)
                elif row_idx in [3, 10, 19, 26, 33]:  # Updated to include new section headers
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 50
        
        # Sort main sheet by Montgomery Score (descending)
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(output_file)
        print(f"\nAnalysis complete! Results saved to: {output_file}")
        
        # Print top stocks
        if not df.empty:
            print("\n" + "="*80)
            top_stocks = df.nlargest(5, 'Montgomery Score')[['Ticker', 'Company Name', 'Montgomery Score', 'Quality Rating', 'ROE (%)', 'Debt to Equity']]
            print("\nTop 5 Stocks by Montgomery Score:")
            print(top_stocks.to_string(index=False))

            # Print top stocks by upside potential
            if 'Upside Potential (%)' in df.columns:
                print("\n" + "="*80)
                # Filter for positive upside and sort
                upside_stocks = df[df['Upside Potential (%)'] > 0].nlargest(5, 'Upside Potential (%)')[
                    ['Ticker', 'Company Name', 'Current Price', 'Avg Intrinsic Value', 'Upside Potential (%)', 'Value Rating', 'Montgomery Score']
                ]
                if not upside_stocks.empty:
                    print("\nTop 5 Stocks by Upside Potential (Highest Payout Opportunity):")
                    print(upside_stocks.to_string(index=False))

            # Print best value opportunities (high upside + high quality score)
            if 'Upside Potential (%)' in df.columns and 'Montgomery Score' in df.columns:
                print("\n" + "="*80)
                # Create a combined value score: stocks with >20% upside and >50 quality score
                value_opps = df[(df['Upside Potential (%)'] > 20) & (df['Montgomery Score'] >= 50)].copy()
                if not value_opps.empty:
                    # Sort by upside potential
                    value_opps = value_opps.nlargest(5, 'Upside Potential (%)')[
                        ['Ticker', 'Company Name', 'Current Price', 'Avg Intrinsic Value', 'Upside Potential (%)', 'Montgomery Score', 'Quality Rating']
                    ]
                    print("\nBest Value Opportunities (High Quality + High Upside):")
                    print(value_opps.to_string(index=False))
                else:
                    print("\nNo stocks found with both high quality (Score ≥50) and high upside (>20%)")
            print("="*80)

def main():
    parser = argparse.ArgumentParser(description='Value Investment Stock Scorer (Montgomery Method)')
    parser.add_argument('ticker_file', help='Text file containing stock tickers (one per line)')
    parser.add_argument('-o', '--output', default='value_investment_analysis.xlsx',
                       help='Output Excel file name (default: value_investment_analysis.xlsx)')
    parser.add_argument('-w', '--workers', type=int, default=1,
                       help='Number of parallel workers for processing (default: 1, sequential)')
    parser.add_argument('--fresh', action='store_true',
                       help='Start fresh analysis, ignoring any existing results file')

    args = parser.parse_args()
    
    # Read tickers from file
    try:
        with open(args.ticker_file, 'r') as f:
            # Filter out empty lines and comments (lines starting with #)
            tickers = [line.strip() for line in f if line.strip() and not line.strip().startswith('#')]
    except FileNotFoundError:
        print(f"Error: File '{args.ticker_file}' not found")
        sys.exit(1)
    
    if not tickers:
        print("Error: No tickers found in file")
        sys.exit(1)
    
    print(f"Found {len(tickers)} tickers in input file")
    print("=" * 60)

    # Process stocks
    scorer = MontgomeryValueScorer()

    # Load existing results if output file exists (for resume functionality)
    # Skip if --fresh flag is used
    if not args.fresh:
        scorer.load_existing_results(args.output)
    else:
        print("Starting fresh analysis (--fresh flag used)")

    print(f"\nAnalyzing stocks using Montgomery Value methodology...")
    print("=" * 60)

    results_df = scorer.process_tickers(tickers, workers=args.workers)

    # Check if we have any new results or existing results
    if results_df.empty and scorer.existing_results is None:
        print("No valid data retrieved for any ticker")
        sys.exit(1)

    if results_df.empty and scorer.existing_results is not None:
        print("\nAll tickers have already been processed!")
        print(f"Total tickers in {args.output}: {len(scorer.existing_results)}")
        sys.exit(0)

    # Sort by Montgomery Score (will be done after merging in create_excel_report)
    if not results_df.empty:
        results_df = results_df.sort_values('Montgomery Score', ascending=False)

    # Create Excel report (this will merge with existing results if any)
    scorer.create_excel_report(results_df, args.output)

if __name__ == "__main__":
    main()
